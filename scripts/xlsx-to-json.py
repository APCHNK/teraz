#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Konwerter "Dr Teraz _ Tabela z ofertą" (xlsx) -> JSON dla importera w motywie
(inc/tz-offer-import.php).

Użycie:
    python3 scripts/xlsx-to-json.py "/ścieżka/do/Dr Teraz _ Tabela z ofertą_v3.xlsx"

Generuje:
    inc/import-data/szczepienia.json   (produkty WooCommerce, klucz: SKU)
    inc/import-data/preparaty.json     (CPT preparat, klucz: tytuł)
    inc/import-data/uslugi.json        (CPT usluga, klucz: tytuł + miasto)
"""
import json
import os
import sys

import openpyxl

CITY_PREFIX = {
    'Poznań': 'PO',
    'Łódź': 'LO',
    'Gdańsk': 'GD',
    'Warszawa': 'WA',
    None: 'OG',
}

OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'inc', 'import-data')


def cell(v):
    """Normalizacja komórki: None/'' -> '', liczby -> str bez .0, trim."""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def rows_as_dicts(ws):
    headers = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None or str(r[0]).strip() == '':
            continue
        yield {headers[i]: r[i] for i in range(len(headers)) if headers[i]}


def convert_szczepienia(wb):
    out = []
    for d in rows_as_dicts(wb['Szczepienia']):
        miasto = cell(d.get('Miasto')) or None
        prefix = CITY_PREFIX.get(miasto)
        if prefix is None:
            raise SystemExit(f"Nieznane miasto w arkuszu Szczepienia: {miasto!r}")
        num = cell(d.get('ID'))
        corrected_sku = f"{prefix}-{num}"
        sheet_sku = cell(d.get('SKU'))
        if sheet_sku != corrected_sku:
            print(f"  UWAGA: SKU w arkuszu {sheet_sku!r} != oczekiwane {corrected_sku!r} "
                  f"({d['Choroba']} / {miasto or 'Ogólne'}) - importer zapisze {corrected_sku!r}")
        kraje = [k.strip() for k in cell(d.get('Kraje-upsell2')).split(',') if k.strip()]
        # Filtry (kolumna "Filtry" to konkatenacja tych czterech kolumn)
        filtry = [cell(d.get(c)) for c in ('Podróżne', 'HPV', 'Dla dzieci', 'Sezonowe') if cell(d.get(c))]
        crosssell = [s.strip() for s in cell(d.get('Cross sell')).split(',') if s.strip()]
        title = f"Szczepienie {cell(d['Choroba'])}" + (f" {miasto}" if miasto else "")
        out.append({
            'sku': corrected_sku,
            'sheet_sku': sheet_sku,
            'id_szczepienia': num,
            'choroba': cell(d['Choroba']),
            'miasto': miasto or '',
            'miasto_slug': cell(d.get('Miasto slug2')),
            'title': title,
            'content': cell(d.get('Opis')),
            'menu_order': cell(d.get('Menu order')),
            'naglowek': cell(d.get('Nagłówek H1')),
            'badanie': cell(d.get('Badanie')),
            'schemat': cell(d.get('Schemat')),
            'liczba_dawek': cell(d.get('Liczba dawek')),
            'rodzaj': cell(d.get('Rodzaj')),
            'droga_zakazenia': cell(d.get('Droga zakażenia')),
            'preparat_google': cell(d.get('Preparat - google')),
            'preparat': cell(d.get('Preparat')),
            'dostepnosc': cell(d.get('Dostępność')),
            'cena': cell(d.get('Cena za 1 dawkę')),
            'czas_do_uodpornienia': cell(d.get('Czas do uodpornienia')),
            'kraje': kraje,
            'filtry': filtry,
            'crosssell_skus': crosssell,
        })
    return out


def convert_preparaty(wb):
    out = []
    for d in rows_as_dicts(wb['Szczepienia preparaty']):
        out.append({
            'title': cell(d['Preparat']),
            'choroby': cell(d.get('Choroby')),
            'id_szcz': cell(d.get('ID')),
            'schemat': cell(d.get('Schemat szczepień')),
            'droga_podania': cell(d.get('Droga podania')),
            'finansowanie_nfz': cell(d.get('Finansowanie NFZ')),
            'typ': cell(d.get('Typ')),
            'dostepnosc': cell(d.get('Dostępność')),
            'ciaza': cell(d.get('Ciąża')),
            'naglowek': cell(d.get('H1')),
            'opis': cell(d.get('Opis')),
            'chpl': cell(d.get('Chpl')),
        })
    return out


def convert_uslugi(wb):
    out = []
    for d in rows_as_dicts(wb['Usługi']):
        miasto = cell(d.get('Miasto')) or 'Ogólne'
        faq = []
        for i in range(1, 7):
            faq.append({
                'enabled': cell(d.get(f'Akordeon {i}')) == 'Tak',
                'tytul': cell(d.get(f'Akordeon {i} tytuł')),
                'tresc': cell(d.get(f'Akordeon {i} treść')),
            })
        kategorie = [k.strip() for k in cell(d.get('Kategoria')).split(',') if k.strip()]
        out.append({
            'title': cell(d['Nazwa']),
            'miasto': miasto,
            'slug_suffix': cell(d.get('Slug')),
            'dostepne_wkrotce': cell(d.get('Dostępne wkrótce')),
            'naglowek_wyswietlany': cell(d.get('Nagłówek')),
            'podtytul': cell(d.get('Podtytuł')),
            'formularz': cell(d.get('Formularz')),
            'opis': cell(d.get('Opis')),
            'rezerwacja': cell(d.get('Rezerwacja')),
            'niemcewicza': cell(d.get('Niemcewicza')),
            'modlinska': cell(d.get('Modlińska')),
            'grabowa': cell(d.get('Poznań Grabowa')),
            'kategorie': kategorie,
            'cena_poz': cell(d.get('Cena POZ')),
            'cena_prywatnie': cell(d.get('Cena prywatnie')),
            'zasady_korzystania': cell(d.get('Zasady korzystania')),
            'archiwum_shortcode': cell(d.get('Shortcode archiwum')),
            'obraz_nad_opisem': cell(d.get('Obraz nad opisem')),
            'faq': faq,
        })
    return out


def main():
    if len(sys.argv) != 2:
        raise SystemExit(__doc__)
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
    os.makedirs(OUT_DIR, exist_ok=True)
    for name, data in (
        ('szczepienia', convert_szczepienia(wb)),
        ('preparaty', convert_preparaty(wb)),
        ('uslugi', convert_uslugi(wb)),
    ):
        path = os.path.join(OUT_DIR, f'{name}.json')
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=1)
        print(f"{name}: {len(data)} rekordów -> {os.path.relpath(path)}")


if __name__ == '__main__':
    main()
